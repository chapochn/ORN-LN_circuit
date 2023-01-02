"""
containing general use functions
@author: Nikolai M Chapochnikov
"""


# #############################################################################
# ################################# IMPORTS ###################################
# #############################################################################
import numpy as np
import openpyxl
import pandas as pd
import itertools
import scipy.linalg as LA
# import os
import pathlib
import datetime
from scipy import stats as SS
from typing import Tuple, Dict, Union



# #############################################################################
# ##########################  FILE MANAGEMENT/READING  ########################
# #############################################################################


def create_path(path):
    now = datetime.datetime.now()
    now_dir = now.strftime("%Y%m%d_%H%M%S/")
    new_path = path / now_dir
    new_path.mkdir()
    # if not os.path.exists(new_path):
    #     os.makedirs(new_path)
    return new_path


def get_str_array_from_h5(h5_data, dat, dtype='U'):
    dat = np.array(dat).flatten()
    for i in range(len(dat)):
        my_str = np.array(h5_data[dat[i]]).flatten()
        my_str = ''.join([chr(c) for c in my_str])
        dat[i] = my_str
    dat = dat.astype(dtype)
    return dat


def get_sheet(file_name, sheet_name=''):
    wb = openpyxl.load_workbook(file_name)
    if sheet_name == '':
        print('sheet names:', wb.sheetnames)
        sheet_name = wb.sheetnames[0]
    return wb[sheet_name]


def get_data(sheet, i_begin, i_n, j_begin, j_n):  # first rows, then columns
    output = np.zeros((i_n, j_n))
    for i in range(i_n):
        for j in range(j_n):
            output[i, j] = sheet.cell(row=i+i_begin, column=j+j_begin).value
    return output


def get_labels_clmn(sheet, i_begin, i_n, j):
    output = []
    for i in range(i_n):
        output.append(sheet.cell(row=i+i_begin, column=j).value)
    return output


def get_labels_row(sheet, i_begin, i_n, j):
    output = []
    for i in range(i_n):
        output.append(sheet.cell(row=j, column=i+i_begin).value)
    return output


# #############################################################################
# ##########################  STRING FUNCTIONS  ###############################
# #############################################################################


def get_strs_in_cats(str_list, cats, compl=False):
    """
    str_list is the general list of strings, usually the list of all cell names
    cats is a selection of strings, it is usually a list
    for any x in str_list, if x is part of any y (y is an element of cats)
    then x is chosen. Example of str_list: all the cells in the experiment
    example of cats: ['Broad T', 'Keystone L'], then all the cells which have
    a substring 'Broad T' or the substring 'Keystone L' will be chosen.
    An other way of putting it: choosing all the cells that are part
    of any of the categories in cats
    the option compl gives the complementary set of names
    """
    if not compl:
        return [name for name in str_list if any(x in name for x in cats)]
    else:
        return [name for name in str_list if not any(x in name for x in cats)]


def get_cat(cell_name, cats):
    """
    return the category to which the cell_name belongs to
    """
    for cat in cats:
        if cat in cell_name:
            return cat
    raise ValueError(f'{cell_name} seem to not belong to any category {cats}')


def replace_substr(string: str, dict_str: Dict[str, str]):
    for key, val in dict_str.items():
        string = string.replace(key, val)
    return string


def replace_substr_np(array: Union[list, np.ndarray],
                      dict_str: Dict[str, str]) -> Union[list, np.ndarray]:
    for i in range(len(array)):
        array[i] = replace_substr(array[i], dict_str)
    return array


def is_str_included(A: Union[list, np.ndarray], B: Union[list, np.ndarray])\
        -> bool:
    """
    Checks if every string in the list A is a substring of B
    """
    return [x for x in A if any(x in y for y in B)] == A


def repl_add(name: str, old: str, new: str):
    """
    if the string old is in the string name, it is removed, and the string new
    is added at the end of the string name
    the updated string is returned
    """
    if old in name:
        name = name.replace(old, '')
        name = name + new
    return name


def repl_preadd(name: str, old: str, new: str):
    """
    if the string old is in the string name, it is removed, and the string new
    is addded at the beginning of the string name
    the updated string is returned
    """
    if old in name:
        name = name.replace(old, '')
        name = new + name
    return name

# =============================================================================
# def get_matching_names(sub_list, full_list):
#     # take a name in the list with the shorter names
#     short_str = sub_list[0]
#     # find in the list with longer names the equivalent name, just longer
#     long_str = list(filter(lambda x: short_str in x, full_list))[0]
#     # find what is missing
#     suffix = long_str[len(short_str):]
#     # add it to the initial list of names
#     chosen_list = [name + suffix for name in sub_list]
#
#     check = np.array([name in full_list for name in chosen_list])
#     if np.sum(check-1) != 0:
#         print('we have a problem in get_matching_names')
#
#     return chosen_list
# =============================================================================


# this version should be more stable. The one above should be removed
def get_match_names(sub_list, full_list):
    chosen_list = []
    # take a name in the list with the shorter names
    for short_str in sub_list:
        # find in the list with longer names the equivalent name, just longer
        long_str = list(filter(lambda x: short_str in x, full_list))[0]
        # add it to the initial list of names
        chosen_list.append(long_str)

    check = np.array([name in full_list for name in chosen_list])
    if np.sum(check-1) != 0:
        print('we have a problem in get_match_names')

    return chosen_list


def fullprint(data):
    opt = np.get_printoptions()
    np.set_printoptions(threshold=np.inf)
    np.set_printoptions(suppress=True)
    np.set_printoptions(linewidth=np.nan)
    print(data)
    np.set_printoptions(**opt)


def have_same_index(A: pd.DataFrame, B: pd.DataFrame) -> bool:
    return set(A.index) == set(B.index)


def align_indices(A: pd.DataFrame, B: pd.DataFrame)\
        -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    What this function does is making the 2 dataframes have the same order
    in the index. And in addition, it also covers the case when the indices
    are not exactly the same (we work with strings here),
    but one name is shorter than the other. In that case strings also become
    the same in both dataframes, and the strings becore as in the dataframe
    with the longer strings.

    the alignement happens according to the first dataset A, so A is not
    changed, but B is sorted

    A and B must be DataFrames, strings become longer
    """
    if len(A) != len(B):
        raise ValueError(f"datasets don't have the same index length: "
                         + f"{A.shape} {B.shape}")

    if set(A.index) == set(B.index):
        pass
    elif is_str_included(list(A.index), list(B.index)):
        l1 = get_match_names(list(A.index), list(B.index))
        dict_cells = dict(zip(list(A.index), l1))
        A = A.rename(index=dict_cells)
    elif is_str_included(list(B.index), list(A.index)):
        l1 = get_match_names(list(B.index), list(A.index))
        dict_cells = dict(zip(list(B.index), l1))
        B = B.rename(index=dict_cells)
    else:
        raise ValueError('Datasets do not have same indices')

    # this is just a double check
    if set(A.index) == set(B.index):
        srt = A.index
        B = B.loc[srt]
    else:
        raise ValueError('Datasets do not have same indices')
    return A, B


def get_change_idx(df: pd.DataFrame, level):
    x = list(df.index.get_level_values(level))
    x_change = np.where(np.diff(x) != 0)[0] + 1
    return x_change


# #############################################################################
# ##########################  CALCULUS FUNCTIONS  #############################
# #############################################################################


def add_point1(X: np.ndarray) -> np.ndarray:
    """
    for 1d arrays, adds a point at the end of the array, which a duplicate
    """
    return np.append(X, X[-1])


def add_point2(X: np.ndarray) -> np.ndarray:
    """
    for 2d arrays
    """
    # here concatenate can be replaced by append without change of results
    X1 = np.array([X[:, -1]]).T
    return np.concatenate((X, X1), axis=1)


def add_point3(X: np.ndarray) -> np.ndarray:
    """
    for 3d arrays
    """
    X1 = np.rollaxis(np.array([X[:, :, -1]]), 0, 3)
    return np.concatenate((X, X1), axis=2)


def rectify(data: Union[np.ndarray, pd.DataFrame])\
        -> Union[np.ndarray, pd.DataFrame]:
    # return data * (np.sign(data)+1.)/2.
    return np.maximum(0, data)  # probably faster...


def subtr_diag(A, doit: bool = True):
    '''
    just removing the diagonal of a square matrix
    '''
    if doit:
        return A - np.diag(np.diag(A))
    else:
        return A


def shift_pos(data):
    return data - np.min(data)


def func_hill(x, y_max, EC_50, n):
    return y_max * x**n/(x**n + EC_50**n)


def get_ctr_norm(X, opt=0):
    """
    for dataframes, each column is centered and normalized
    for series, the serie is centered and normalized
    """
    if X.ndim == 1:
        return get_ctr_norm1(X, opt)
    elif X.ndim == 2:
        return get_ctr_norm2(X, opt)
    else:
        raise TypeError('there is something wrong with the dataset')


def get_ctr_norm1(X, opt=0):
    """
    center and normalization for series
    options: 0, 1, 2
    0: returns cn
    1: returns (c, cn)
    2: returns (c, n, cn)
    """
    X_c = X - X.mean()
    X_cn = X_c / LA.norm(X_c)
    if opt == 0:
        return X_cn
    elif opt == 1:
        return X_c, X_cn
    elif opt == 2:
        X_n = X / LA.norm(X)
        return X_c, X_n, X_cn
    else:
        raise ValueError('there is no option ' + str(opt))


def get_ctr_norm2(X, opt=0):
    """
    each column is centered and normalized
    opt can be 0, 1, 2
    0: returns cn
    1: returns (c, cn)
    2: returns (c, n, cn)
    """
    X_c = X.subtract(X.mean(axis=0))
    X_cn = X_c.divide(LA.norm(X_c, axis=0), axis=1)
    if opt == 0:
        return X_cn
    elif opt == 1:
        return X_c, X_cn
    elif opt == 2:
        X_n = X.divide(LA.norm(X, axis=0), axis=1)
        return X_c, X_n, X_cn
    else:
        raise ValueError('there is no option ' + str(opt))


def get_ctr_norm_np(X, opt=0):
    """
    for dataframes, each columns is centered and normalized
    for series, the serie is centered and normalized
    """
    if X.ndim == 1:
        return get_ctr_norm1(X, opt)
    elif X.ndim == 2:
        return get_ctr_norm2_np(X, opt)
    else:
        raise TypeError('there is something wrong with the dataset')


def get_ctr_norm2_np(X, opt=0):
    """
    each column is centered and normalized, which is the 2nd dimension
    Meaning that the 2nd index is kept fix, and the first index is the one that
    is the different coordinate/indexes of the vector
    opt can be 0, 1, 2
    0: returns cn
    1: returns (c, cn)
    2: returns (c, n, cn)
    """
    X_c = X - X.mean(axis=0)
    X_cn = X_c / LA.norm(X_c, axis=0)
    if opt == 0:
        return X_cn
    elif opt == 1:
        return X_c, X_cn
    elif opt == 2:
        X_n = X / LA.norm(X, axis=0)
        return X_c, X_n, X_cn
    else:
        raise ValueError('there is no option ' + str(opt))


def get_norm(X):
    """
    for dataframes, each columns is normalized
    for series, the serie is centered and normalized
    """
    if X.ndim == 1:
        return get_norm_np(X)
    elif X.ndim == 2:
        return get_norm2(X)
    else:
        raise TypeError('there is something wrong with the dataset')


def get_norm2(X):
    """
    each column is normalized
    """
    return X.divide(LA.norm(X, axis=0), axis=1)


def get_norm_np(X):
    """
    for dataframes, each columns is normalized
    for series, the serie is normalized
    """
    return X / LA.norm(X, axis=0)


def is_mean_0(A, verbose: bool = False) -> bool:
    """
    tests if the mean is 0
    might be useful to rewrite using the function allclose
    """
    A_c = A.mean()
    A_max = np.abs(A).max(skipna=False)
    if np.sum(A_max) == 0:
        raise ValueError('The data is not clean')

    zeros = (A_max == 0) | np.isnan(A_max)
    # this is the number of places where it is not 0
    non_zeros1 = ((A.loc[:, zeros].sum() == 0) - 1).sum()
    non_zeros2 = ((A_c[~zeros]/A_max[~zeros] < 1e-15) - 1).sum()
    non_zeros3 = ((A_c[~zeros] < 1e-15) - 1).sum()
    if verbose is True:
        print(non_zeros1, non_zeros2, non_zeros3)
    cond1 = (non_zeros1 + non_zeros2 == 0)
    cond2 = (non_zeros1 + non_zeros3 == 0)
    return cond1 or cond2


def is_norm_1(A: pd.DataFrame, verbose: bool = False) -> bool:
    """
    testing if the norm is 1, A is a pandas array here
    probably should write a similar function for a pure numpy array
    also i imagine there is a better way to test this, maybe something is
    already implemented of that sort
    YES: would make sense to rewrite with the function allclose
    """
    A_max = np.abs(A).max(skipna=False)
    if np.sum(A_max) == 0:
        raise ValueError('the data is not clean')

    zeros = (A_max == 0) | np.isnan(A_max)
    norms = LA.norm(A.loc[:, ~zeros], axis=0)
    norms_1 = (np.abs((norms - 1)) < 1e-15)
    norms_cond = (norms_1 - 1).sum()  # number of mismatches
    if verbose is True:
        print(norms_cond)
    return norms_cond == 0


def is_norm_1_np(A: np.ndarray, verbose: bool = False) -> bool:
    """
    testing if the norm is 1, A is a pandas array here
    probably should write a similar function for a pure numpy array
    also i imagine there is a better way to test this, maybe something is
    already implemented of that sort
    YES: would make sense to rewrite with the function allclose
    """
    A_max = np.abs(A).max()
    if np.sum(A_max) == 0:
        raise ValueError('the data is not clean')

    zeros = (A_max == 0) | np.isnan(A_max)
    norms = LA.norm(A[:, ~zeros], axis=0)
    norms_1 = (np.abs((norms - 1)) < 1e-15)
    norms_cond = (norms_1 - 1).sum()  # number of mismatches
    if verbose is True:
        print(norms_cond)
    return norms_cond == 0


def get_proj_mat(v: Union[pd.DataFrame, np.ndarray]) -> np.ndarray:
    """
    gives the projection matrix based of the given column vectors
    Parameters
    ----------
    v

    Returns
    -------

    """
    if isinstance(v, pd.DataFrame):
        v = v.values

    inv = LA.inv(v.T @ v)
    return v @ inv @ v.T


def get_pdf_1(data, bins_pdf, add_point=True, cdf_bool=False, checknan=False):
    """
    data is a 1d array
    the function gives back the pdf
    """
# =============================================================================
#     pdf, _ = np.histogram(corr, bins=len(bins_pdf)-1,
#                           range=(bins_pdf[0], bins_pdf[-1]))
# =============================================================================

    if checknan and np.isnan(data).any():  # if there is a single nan there
        pdf = np.full(len(bins_pdf) - 1, np.nan)
    else:
        pdf, _ = np.histogram(data, bins=bins_pdf)
        if cdf_bool:
            pdf = np.cumsum(pdf)/len(data)

    # adding a duplicated data point useful for the plotting:
    if add_point:
        pdf = add_point1(pdf)

    return pdf


def get_pdf_2(data, bins_pdf, add_point=True, cdf_bool=True, checknan=False):
    """
    data is a 2d array, the first dimension are the iterations
    the function gives back the pdfs
    """
    N_iter = len(data)
    pdfs = np.zeros((N_iter, len(bins_pdf) - 1))
    for i in range(N_iter):
        pdfs[i] = get_pdf_1(data[i], bins_pdf, add_point=False,
                            cdf_bool=False, checknan=checknan)

    if cdf_bool:
        pdfs = np.cumsum(pdfs, axis=1)/data.shape[1]

    if add_point:
        pdfs = add_point2(pdfs)

    return pdfs


def get_pdf_3(data, bins_pdf, add_point=True, cdf_bool=False,
              checknan=False):
    """
    data is a 3d array, the first dimension are the iterations, the second
    dimension is usually the cells
    the function gives back the pdf

    add_point option duplicated the last point
    checknan checks if there are any nans in the set and gives nan as
    result for the pdf instead 0 as would be calculated naturally
    """
    N1, N2, N3 = data.shape
    pdfs = np.zeros((N1, N2, len(bins_pdf) - 1))
    for i in range(N1):
        pdfs[i]  = get_pdf_2(data[i], bins_pdf, add_point=False,
                             cdf_bool=False, checknan=checknan)

    if cdf_bool:
        pdfs = np.cumsum(pdfs, axis=2)/data.shape[2]

    if add_point:
        pdfs = add_point3(pdfs)

    return pdfs


def get_pdf_cdf_1(corr, bins_pdf, bins_cdf, add_point=True, cdf_bool=True,
                  checknan=False):
    
    """
    corr is a 1d array
    the function gives back the pdf and the cdf
    """
# =============================================================================
#     pdf, _ = np.histogram(corr, bins=len(bins_pdf)-1,
#                           range=(bins_pdf[0], bins_pdf[-1]))
# =============================================================================

    if checknan and np.isnan(corr).any():  # if there is a single nan there
        pdf = np.full(len(bins_pdf) - 1, np.nan)
        cdf = np.full(len(bins_cdf) - 1, np.nan)
    else:
        pdf, _ = np.histogram(corr, bins=bins_pdf)
        cdf, _ = np.histogram(corr, bins=bins_cdf)
        if cdf_bool:
            cdf = np.cumsum(cdf)/len(corr)

    # cumfreq is much slower (about 4 folds) because it is calculating the
    # linspace at each call
# =============================================================================
#     cdf, _, _, _ = SS.cumfreq(corr, numbins=len(bins_cdf)-1,
#                               defaultreallimits=(bins_cdf[0], bins_cdf[-1]))
#     cdf /= len(corr)
# =============================================================================
    # adding a duplicated data point useful for the plotting:
    if add_point:
        pdf = add_point1(pdf)
        cdf = add_point1(cdf)

    return pdf, cdf


def get_pdf_cdf_2(corr, bins_pdf, bins_cdf, add_point=True, cdf_bool=True,
                  checknan=False):
    """
    corr is a 2d array, the first dimension are the iterations
    the function gives back the pdfs and the cdfs
    """
    N_iter = len(corr)
    pdfs = np.zeros((N_iter, len(bins_pdf) - 1))
    cdfs = np.zeros((N_iter, len(bins_cdf) - 1))
    for i in range(N_iter):
        pdfs[i], cdfs[i] = get_pdf_cdf_1(corr[i], bins_pdf, bins_cdf,
                                         add_point=False, cdf_bool=False,
                                         checknan=checknan)

    if cdf_bool:
        cdfs = np.cumsum(cdfs, axis=1)/corr.shape[1]

    if add_point:
        pdfs = add_point2(pdfs)
        cdfs = add_point2(cdfs)

    return pdfs, cdfs


def get_pdf_cdf_3(corr, bins_pdf, bins_cdf, add_point=True, cdf_bool=True,
                  checknan=False):
    """
    corr is a 3d array, the first dimension are the iterations, the second
    dimension is usually the cells
    the function gives back the pdf and the cdf

    add_point option duplicated the last point
    checknan checks if there are any nans in the set and gives nan as
    result for the pdf and cdf instead 0 as would be calculated naturally
    """
    N1, N2, N3 = corr.shape
    pdfs = np.zeros((N1, N2, len(bins_pdf) - 1))
    cdfs = np.zeros((N1, N2, len(bins_cdf) - 1))
    for i in range(N1):
        pdfs[i], cdfs[i] = get_pdf_cdf_2(corr[i], bins_pdf, bins_cdf,
                                         add_point=False, cdf_bool=False,
                                         checknan=checknan)

    if cdf_bool:
        cdfs = np.cumsum(cdfs, axis=2)/corr.shape[2]

    if add_point:
        pdfs = add_point3(pdfs)
        cdfs = add_point3(cdfs)

    return pdfs, cdfs


def get_max_diff(curves, curve1):
    """
    curves can be 1D, 2D, 3D.
    the max is always done on the last dimension which is the dimension
    of one individual curve
    curve1 is (are) the mean curve(s)
    """
    diffs = curves - curve1
    return np.max(diffs, axis=-1)


def get_min_diff(curves, curve1):
    """
    curves can be 1D, 2D, 3D.
    the max is always done on the last dimension which is the dimension
    of one individual curve
    curve1 is (are) the mean curve(s)
    returning the absolute deviation
    """
    diffs = curves - curve1
    return -np.min(diffs, axis=-1)


def get_andersondarling_dist(curves, curve1):
    """
    curves can be 1D, 2D, 3D.
    the max is always done on the last dimension which is the dimension
    of one individual curve
    curve1 is (are) the mean curve(s)
    returning Anderson Darling statistic
    """
    diffs = curves - curve1
    dF = np.diff(curve1, axis=-1, append=1)/ (curve1 * (1-curve1))
    dF[~np.isfinite(dF)] = 1
    dF[dF==0] = 1
    return np.sum(diffs**2 * dF , axis=-1)


def get_entries(A: Union[pd.DataFrame, np.ndarray], diag: bool=False)\
        -> np.ndarray:
    """
    need to be a square matrix
    returns the entries of the matrix, ignoring the diagonal elements, apart
    if diag=True
    """
    if isinstance(A, pd.DataFrame):
        A = A.values

    if diag:
        return A.flatten()
    else:
        idx = ~np.eye(len(A), dtype=bool)
        return A[idx]


def is_permutation_matrix(x):
    """[summary]
    
    Arguments:
        x {[type]} -- [description]
    
    Returns:
        [type] -- [description]
    """
    x = np.asanyarray(x)
    return (x.ndim == 2 and x.shape[0] == x.shape[1] and
            (x.sum(axis=0) == 1).all() and
            (x.sum(axis=1) == 1).all() and
            ((x == 1) | (x == 0)).all())


FLOAT_TYPE = np.float32
FLOAT_TYPE = np.float64
EPSILON = np.finfo(FLOAT_TYPE).eps


def mat_mul1(A, B, alpha=1):
    """
    https://www.benjaminjohnston.com.au/matmul
    """
    # return np.matmul(A, B)
    return LA.blas.sgemm(alpha, A, B)



def mat_mul2(A, B, alpha=1):
    """
    https://www.benjaminjohnston.com.au/matmul
    """
    return alpha * np.matmul(A, B)
    # return LA.blas.sgemm(alpha, A, B)


if FLOAT_TYPE == np.float32:
    mat_mul = mat_mul1
else:
    mat_mul = mat_mul2


def mv_mul(A, x, alpha=1):
    """
    https://www.benjaminjohnston.com.au/matmul
    """
    # return np.matmul(A, B)
    return LA.blas.sgemv(alpha, A, x)


def mat_mul_s(A, B, alpha=1):
    """
    matrix multiplication where A is a symmetric matrix
    i don't see a difference in performance though, maybe didn't try
    for matrices large enough
    https://www.benjaminjohnston.com.au/matmul
    """
    # return np.matmul(A, B)
    # return LA.blas.ssymm(alpha, A, B, side=1, c=B, beta=1)
    return LA.blas.ssymm(alpha, A, B, side=1)


def get_random_spd_matrix(n_dim: int, eig_vals: np.ndarray) -> np.ndarray:
    """
    generates a random symmetric positive definite matrix of size n_dim
    and with eig_values given
    """
    U = SS.ortho_group.rvs(n_dim)
    return U.T @ np.diag(eig_vals) @ U
    # other options:
    # sklearn.datasets.make_spd_matrix(n_dim, random_state=None)
    # using A, a random matrix and A.T @ A


# #############################################################################
# ################## ACTIVITY SVD ANALYSIS FUNCTIONS ##########################
# #############################################################################
def get_pps(df, pps=None):
    """
    return the origianl data, the centered, the normalized and the centered
    and normalzed version of the original data, put in a dictionary, with
    """
    df_c, df_n, df_cn = get_ctr_norm(df, opt=2)
    df_r = rectify(df)
    to_return = {'o': df, 'c': df_c, 'n': df_n, 'cn': df_cn, 'r': df_r}
    if pps is None:
        pass
    else:
        to_return = {k: v for k, v in to_return.items() if k in pps}
    return to_return


def get_svd_df(act_df, center=False):
    """
    this function should returns the SVD, keeping the labels
    the elements of the dictionary are U, s and Vh
    """
    if center is True:
        act_df = act_df.subtract(act_df.mean(axis=1), axis=0)

    U, s, Vh = LA.svd(act_df, full_matrices=False)
    n = len(s)
    # now we want to create the DataFrames that will keep the labels
    U_df = pd.DataFrame(U, index=act_df.index, columns=np.arange(1, n+1))
    U_df.columns.name = 'PC'
    Vh_df = pd.DataFrame(Vh, columns=act_df.columns, index=np.arange(1, n+1))
    Vh_df.index.name = 'PC'
    s_df = pd.DataFrame(np.diag(s),
                        columns=np.arange(1, n+1), index=np.arange(1, n+1))
    return {'U': U_df, 's': s_df, 'Vh': Vh_df}


def sort_by_pc(U, level):
    """
    This function takes as input the left or right eigenvectors of the
    eigendecomposition of some matrix
    When there are several levels in the MultiIndex, it averages over the
    non-explicited levels.
    """
    if level in U.index.names:
        U = U.T
    elif level not in U.columns.names:
        print("there is a problem sorting")
        return (None, None)

    if U.columns.nlevels == 2:
        sort1 = U.mean(axis=1, level=level).iloc[0].sort_values().index
        sort2 = U.mean(axis=1, level=level).iloc[1].sort_values().index
    elif U.columns.nlevels == 1 and U.columns.name == level:
        sort1 = U.iloc[0].sort_values().index
        sort2 = U.iloc[1].sort_values().index
    else:
        print('there is a problem sorting 2')
        (sort1, sort2) = (None, None)
    return {1: sort1, 2: sort2}


# #############################################################################
# ###############  FUNCTIONS CALCULATING CORR AND SIGNIFICANCE  ###############
# #############################################################################
# it is still not clear to me why i am not not just shuffling the whole
# 2d matrix, it seems it would be much easier and much quicker
# i don't understand why i am shuffling each column one by one.

def get_corr(A, B, check=True):
    """
    getting the correlation coefficient between each columns of dataframe A
    with each column of dataframe B
    First checking if the labels of the rows are aligned and if each
    column is centered and normalized
    A and B should have the same x (row) indices as B, but they can be ordered
    in a different way, if check is true it is first aligning the rows
    of A and B
    in the returned matrix, the rows are the columns of A and the columns
    are the columns of B
    """

    if check:
        # aligning the 2 datasets
        (A, B) = align_indices(A, B)

    if check and ((not is_mean_0(A)) or (not is_mean_0(B))):
        raise ValueError('dataset(s) not centralized: \n'
            f'{np.abs(np.sum(A)).idxmax()}, {np.abs(np.sum(A)).max()}\n'
            f'{np.abs(np.sum(B)).idxmax()}, {np.abs(np.sum(B)).max()}')

    if check and ((not is_norm_1(A)) or (not is_norm_1(B))):
        raise ValueError('dataset(s) not normalized:\n'
                         f'{np.abs(LA.norm(A, axis=0)-1)}, '
                         f'{np.abs(LA.norm(B, axis=0)-1)}')

    # now, the data is aligned, centered and normalized, we know
    # we can calculate the correlation coefficients
    return A.T.dot(B)
    # return A.values.T @ B.values


def get_CS(A, B, check=True):
    """

    getting the Cosine Similarity between each column of dataframe A
    with each column of dataframe B.
    We are only considering positive angles, i.e., between 0 and 1
    angles smaller than 0 are switched sign
    First checking if the labels of the rows are aligned and if each
    column is normalized
    """
    if check:
        # aligning the 2 datasets
        (A, B) = align_indices(A, B)

    if check and ((not is_norm_1(A)) or (not is_norm_1(B))):
        raise ValueError('dataset(s) not normalized:\n'
                         f'{np.abs(LA.norm(A, axis=0)-1)}, '
                         f'{np.abs(LA.norm(B, axis=0)-1)}')

    # now, the data is aligned and normalized, we know
    # we can calculate the cos similarity by
    # return np.abs(A.T.dot(B))
    # wondering if one should or not take the abs here
    return A.T.dot(B)


# some comments on the significance testing:
# in the general case you will have a collection of column-vectors in matrix A
# and a collection of column vectors in matrix B. You are interested in the
# correlation coefficient between each of these vectors. So we are intested
# in correlations all against all.
# Now, when we are doing the significance testing, we are shuffling
# the vector entries.
# in the optimal case we will calculate the shuffling for each vector
# separately. However, we can make things faster


# I think the next function is only used in some old files
# so i am commenting it out
# =============================================================================
# def test_signif_1(A, B, N=10, verbose=False):
#     """
#     we assume that the given vectors are already aligned,
#     centered and normalized
#     for when A is 1D and B is 1D
#     """
#     corr = A @ B
#     corr_shuffled = np.zeros(N)
#
#     for i in range(N):
#         B1 = np.random.permutation(B)
#         corr_shuffled[i] = A @ B1
#
#     if verbose is True:
#         plt.hist(corr_shuffled, 100)
#
#     prob_r = np.sum(corr_shuffled > abs(corr))/N
#     prob_l = np.sum(corr_shuffled < -abs(corr))/N
#     return (corr, np.mean(corr_shuffled), prob_l, prob_r)
# =============================================================================

# almost similar functions to above but for cos similarity
# CS stands for cos-similarity
def get_signif_general(A, B, func, N=10):
    """
    we assume that the given vectors are already aligned and normalized

    A is 2D and B is 1D
    A is d x n
    B is d

    corr, prob_l, prob_r are n

    the option dist returns the cosine similarities that are issued
    from the shuffling

    func is usually either get_corr or get_CS, that's why it is called
    get_signif_general

    return 3 types of pvalues:
    prob_o, it is the right one-tailed pvalue
    prob_r, it is the prob of value coming from random generation comes
    is on the right of abs(real_value)
    prob_l, it is the prob of value coming from random generation comes
    is on the left of -abs(real_value)
    """
    measure = func(A, B, check=False)
    # CS is a 1D vector, so it is the same as abs(A.T @ B)
    # the length of the vector is the same as the number of columns of A, n
    measure_shuffled = np.zeros((N, len(A.T)))  # N x n
    for i in range(N):
        B1 = np.random.permutation(B)
        measure_shuffled[i] = func(A, B1, check=False)

    # we need to look at both the right and left probabilities,
    # as in the case of SVD, the vectors could be oriented in both
    # directions
    prob_o = np.mean(measure_shuffled >= measure, axis=0)  # o - original: 1 tailed
    prob_r = np.mean(measure_shuffled >= np.abs(measure), axis=0)
    prob_l = np.mean(measure_shuffled <= -np.abs(measure), axis=0)
    return measure, measure_shuffled, np.mean(measure_shuffled, axis=0),\
           prob_o, prob_l, prob_r


def get_signif_general_v1(A, B, measure_func, N=10, dist=False):
    """
    this version of significance testing is slower than v2, it shuffles each column
    of B.
    It would make sense to choose as B the data that has less columns,
    so that the shuffling procedure would be faster

    the option dist returns the correlation coefficients that are issued
    from the shuffling, i.e., the full distribution

    measure_func is usually either get_corr or get_CS
    """
    # aligning the 2 datasets, the order in B is kept, A is reordered
    A, B = align_indices(A, B)
    measure_df = measure_func(A, B)
    # the corr_df will have as rows the columns of A and
    # as columns the columns of B

    # now that we have the actual correlation coefficient, we can
    # calculate the distributions of correlations coefficients
    # when the connectivity is shuffled

    A_np = A.values
    B_np = B.values

    if dist:
        measure_collection = np.zeros((N, *measure_df.shape))
    pv_o_df = pd.DataFrame().reindex_like(measure_df)
    pv_l_df = pd.DataFrame().reindex_like(measure_df)
    pv_r_df = pd.DataFrame().reindex_like(measure_df)


    for i in range(len(B.columns)):  # iterating over the columns of B, i.e., each cell
        res = get_signif_general(A_np, B_np[:, i], measure_func, N)
        if dist:
            measure_collection[:, :, i] = res[1]

        pv_o_df.iloc[:, i] = res[3]
        pv_l_df.iloc[:, i] = res[4]
        pv_r_df.iloc[:, i] = res[5]

    if dist:
        return measure_df, measure_collection, pv_o_df, pv_l_df, pv_r_df
    return measure_df, pv_o_df, pv_l_df, pv_r_df

#
# def get_signif_general_v2(A, B, func, N=10, dist=False):
#     """
#     this version of significance testing is much faster than the v1
#     but it might bring some bias.
#     It is basically shuffling all the A matrix at once, and not
#     columns by column as in the test_signif_v1 version
#
#     the option dist returns the correlation coefficients that are issued
#     from the shuffling, i.e., the full distribution
#
#     theoretically we should have that
#     pv_o = combine_pval(CS, pv_l, pv_r)
#     however it is not always exactly the case when CS is negative
#     because in that case in the combine pval it is
#     pval_l that is taken, but pval_l was calculated as
#     np.mean(CS_collection <= -np.abs(CS_df.values), axis=0)
#     that means if certain values in CS_collections are exactly equal
#     (happens in the case when the A or B have some sparcity)
#     and so the CS will be assigned as significative when they are not
#     In any case pv_o is more correct than the combine and should
#     always be preferred.
#     """
#     # aligning the 2 datasets
#     (A, B) = align_indices(A, B)
#     measure_df = func(A, B, check=True)
#
#     A_np = A.values
#     B_np = B.values
#     # measure_collection is a 3D matrix, the first dim are the repetitions
#     # from shuffling, the next 2 dims are the same as for the real CS.
#     measure_collection = np.zeros((N, *measure_df.shape))
#     for i in range(N):
#         measure_collection[i] = func(A_np, np.random.permutation(B_np),
#                                 check=False)
#
#     if dist:
#         return measure_df, measure_collection
#
#     pv_o_df = pd.DataFrame().reindex_like(measure_df)
#     pv_l_df = pd.DataFrame().reindex_like(measure_df)
#     pv_r_df = pd.DataFrame().reindex_like(measure_df)
#     pv = np.mean(measure_collection >= measure_df.values, axis=0)
#     pv_o_df[:] = pv
#     pv = np.mean(measure_collection >= np.abs(measure_df.values), axis=0)
#     pv_r_df[:] = pv
#     pv = np.mean(measure_collection <= -np.abs(measure_df.values), axis=0)
#     pv_l_df[:] = pv
#
#     # we put nan everywhere where there was nans in the corr matrix
#     idx_nan = pd.isnull(measure_df)
#     pv_o_df[idx_nan] = np.nan
#     pv_l_df[idx_nan] = np.nan
#     pv_r_df[idx_nan] = np.nan
#
#     return measure_df, pv_o_df, pv_l_df, pv_r_df


def get_signif_v1(A, B, N=10, dist=False, measure='corr'):
    if measure == 'corr':
        return get_signif_general_v1(A, B, get_corr, N, dist=dist)
    elif measure == 'CS':
        return get_signif_general_v1(A, B, get_CS, N, dist=dist)
    else:
        raise ValueError('no such measure', measure)


# not anymore maintained as not used
# def get_signif_v2(A, B, N=10, dist=False, measure='corr'):
#     # calls function that shuffles all of B simultaneously,
#     # so might add some bias
#     if measure == 'corr':
#         return get_signif_general_v2(A, B, get_corr, N=N, dist=dist)
#     elif measure == 'CS':
#         return get_signif_general_v2(A, B, get_CS, N=N, dist=dist)
#     else:
#         raise ValueError('no such measure', measure)


# # not maintained anymore as replaced by below
# def combine_pval(corrs, pval_l, pval_r):
#     """
#     this function handle the 1-sided vs 2-sided issue of the significance
#     testing
#     here we need to take into account the correlation coefficient
#     (if it is positive or negative), and about the circ_alg we are using
#     if it is SVD, we need we will take the sum of left and right
#     if it is any other circ_alg, it is one-directional
#
#     We rely on the fact that some act processing come from SVD
#     """
#     if pval_l.shape != pval_r.shape:
#         raise ValueError('left and right PV do not have the same shape')
#     if pval_l.shape != corrs.shape:
#         raise ValueError('left sign and corr do not have the same shape')
#
#     # also important to check if the 3 dataframes have the same
#     # index and columns
#     cond1 = pval_l.index.equals(pval_r.index)
#     cond2 = pval_l.index.equals(corrs.index)
#     cond3 = pval_l.columns.equals(pval_r.columns)
#     cond4 = pval_l.columns.equals(corrs.columns)
#     if not (cond1 and cond2 and cond3 and cond4):
#         raise ValueError('pval and/or corr dont have the same index/col')
#
#     (n1, n2) = pval_l.shape
#
#     # final significance to be outputted
#     pval = pd.DataFrame().reindex_like(pval_l)
#
# # could be even faster by creating a function and using apply
#     for i in range(n1):
#         name = corrs.iloc[i].name
#         cond = ('SVD' in name)  # and name[5] != 1)
#         for j in range(n2):
#             corr = corrs.iat[i, j]
#             # in the case of SVD, and if it is not the 1st component
#             # which is expected to be positive
#             if cond:
#                 pv = pval_r.iat[i, j] + pval_l.iat[i, j]
#             else:
#                 # then we are considering cases where we are assuming
#                 # a one sided hypothesis, meaning that if the correlation
#                 # is negative, it will be a very high prob to get
#                 # at least this corr
#                 # this also takes into account the case of the
#                 # 1st loading vector of SVD
#                 if corr >= 0:
#                     pv = pval_r.iat[i, j]
#                 elif corr < 0:
#                     # this part might not be 100% accurate because of the
#                     # of the >= in the signif test, so the values that
#                     # are equal might not be accounted correctly
#                     pv = 1 - pval_l.iat[i, j]
#             pval.iat[i, j] = pv
#     return pval


def combine_pval2(pval_o, pval_l, pval_r):
    """
    this function handle the 1-sided vs 2-sided issue of the significance
    testing
    here we need to take into account the correlation coefficient
    (if it is positive or negative), and about the circ_alg we are using
    if it is SVD, we need we will take the sum
    if it is any other circ_alg, it is one-directional

    We rely on the fact that some act processing come from SVD
    """
    if pval_l.shape != pval_r.shape:
        raise ValueError('left and right sign. do not have the same shape')
    if pval_l.shape != pval_o.shape:
        raise ValueError('left sign and corr. do not have the same shape')

    # also important to check if the 3 dataframes have the same index and cols
    cond1 = pval_l.index.equals(pval_r.index)
    cond2 = pval_l.index.equals(pval_o.index)
    cond3 = pval_l.columns.equals(pval_r.columns)
    cond4 = pval_l.columns.equals(pval_o.columns)
    if not (cond1 and cond2 and cond3 and cond4):
        raise ValueError('pvals do not have the same index/col')

    (n1, n2) = pval_l.shape

    # final significance to be outputted
    pval = pd.DataFrame().reindex_like(pval_l)

    for i in range(n1):
        name = pval_o.iloc[i].name
        cond = ('SVD' in name)  # and name[5] != 1)
        if cond:
            pval.iloc[i] = pval_r.iloc[i] + pval_l.iloc[i]
        else:
            pval.iloc[i] = pval_o.iloc[i]
    return pval


def combine_signif_2tail(corrs, pval_l, pval_r):
    """
    this is the for case when both negative and positive extremes are counted
    as potentially significant
    "https://stats.stackexchange.com/questions/140107/
    p-value-in-a-two-tail-test-with-asymmetric-null-distribution"
    """
    # first some tests to check that the shapes and indeces conform
    if pval_l.shape != pval_r.shape:
        raise ValueError('left and right pval. do not have the same shape')
    if pval_l.shape != corrs.shape:
        raise ValueError('left pval and corr. do not have the same shape')

    # also important to check if the 3 dataframes have the same index and cols
    cond1 = pval_l.index.equals(pval_r.index)
    cond2 = pval_l.index.equals(corrs.index)
    cond3 = pval_l.columns.equals(pval_r.columns)
    cond4 = pval_l.columns.equals(corrs.columns)
    if not (cond1 and cond2 and cond3 and cond4):
        raise ValueError('pval and/or corr do not have the same index/col')

    pval_f = pd.DataFrame().reindex_like(pval_l)

    idx_cor_p = corrs >= 0
    idx_cor_n = np.logical_not(idx_cor_p)

    # the following minimum is necessary for the following condition:
    # imagine the corr is 0.2, and the distribution of corr coef
    # coming from shuffling is centered at + 0.9, so the value 0.2 is
    # an extreme for this distribution. Since in that case the the corr coef
    # is positive, pval_r is the appropriate pval, howwever pval is larger
    # than 1 - pval
    pval_f[idx_cor_p] = np.minimum(pval_r, 1-pval_r)
    pval_f[idx_cor_n] = np.minimum(pval_l, 1-pval_l)

    return 2 * pval_f


def shuffle_matrix(A: np.ndarray) -> np.ndarray:
    """
    shuffles each column of the matrix independently
    Parameters
    ----------
        A: np.ndarray
    """
    A_shfl = A.copy()
    for i in range(len(A.T)):
        A_shfl[:, i] = np.random.permutation(A_shfl[:, i])
    return A_shfl

# #############################################################################
# ##################  SIGNIFICANCE TESTS FOR RAW ORN RESPONSES  ###############
# #############################################################################


def get_mean_signif1(corr, corr_col, fun=np.mean):
    """
    getting the mean correlation and the pvalue for when corr is a 1D and
    corr_col is 2D
    corr are the real correlations and corr_col and the corr resulting from
    shuffling or some other random generation
    instead of the mean, one possibility is to use the median for example
    """
    N_iter, N_odors = corr_col.shape
    corr_col_means = fun(corr_col, axis=1)
    corr_m = fun(corr)  # there shouldn't be any nan here
    # if there is a nan present in the next line, it will give false,
    # which i think is what we want. Because if for a certain configuration
    # the corrcoef is not defined, then it makes sense that that configuration
    # is not "better" than the control corr coef

    pval_r = np.mean(corr_col_means >= corr_m)  # r as right

    if pval_r == 0:
        pval_r = 1./N_iter
    pval2 = 2 * np.min((pval_r, 1-pval_r))
    return corr_m, pval_r, pval2


def get_mean_signif2(corr: pd.DataFrame, corr_col: pd.DataFrame):
    """
    getting the mean correlation and the pvalue for when corr is a 2D and
    corr_col is 3D
    corr is n_cells x n_odors
    corr_col is N_iter x n_cells x n_odors
    corr are the real correlations and corr_col and the corr resulting
    from shuffling or some other random generation

    pval is the right 1-tailed
    pval2 is the 2-tailed pvalue
    """
    N_iter, N_cells, N_odors = corr_col.shape
    corr_col_m = np.mean(corr_col, axis=2)  # mean over all odors
    corr_m = np.mean(corr, axis=1)
    pval = np.mean(corr_col_m >= corr_m.values, axis=0)
    pval = pd.Series(pval, index=corr.index)
    pval = pval.replace(0, 1./N_iter)

    pval2 = 2 * np.minimum(pval, 1 - pval)

    return corr_m, pval, pval2

# not maintained anymore
# def characterize_raw_responses(act, con, N1, N2):
#     """
#     we expect that the act and con vectors are already centered and
#     normalized to make all the necessary calculations to get the corr
#     and signif
#
#     N1 corresponds to the number of iterations to calculate the significance
#     per odors and per cll
#
#     N2 corresponds to the number of iterations to create fake cdfs and thus
#     calculate the p-values for each cell.
#
#     what the function returns:
#     corr: matrix of correlation coefficients for each cell and simulus
#     corr_pval: matrix of pvalues
#     corr_m_per_cell: mean corr coef per cell
#     corr_m_per_cell_pval: pval based on the mean corr coef per cell
#     cdf_diff_max: maximum deviation in the cdf per cell
#     cdf_diff_max_pval: pvalue based on the max deviation in the cdf per cell
#     cdf_true: actual cdf per cell
#     cdfs_shfl: all the cdfs that were created on the way
#     """
#
#     cell_list = list(con.columns)
#
#     # getting the odors vs con correlations and pvals
#     corr, pval_o, _, _ = get_signif_corr_v2(act, con, N=N1)
#     corr = corr.T
#     # when we are combining the correlation coefficients in a way that
#     # only it is a low pvalue if the corr is very high, and it will be a large
#     # pvalue if the corr coef is small
#     corr_pval = pval_o.T  # combine_pval(corr, pval_l.T, pval_r.T)
#     # print(np.max(np.abs((corr_pval - pval_o.T).values)))
#     corr_pval = corr_pval.replace(0, 1./N1)
#
#     # corr is n_cells x n_odors
#     # corr_col is N1 x n_cells x n_odors
#     _, corr_col = get_signif_corr_v2(act, con, N=N2, dist=True)
#     corr_col = np.swapaxes(corr_col, 1, 2)
#     if np.isnan(corr_col).any():
#         print("WE HAVE SOME NANs IN corr_col which shouldn't be there")
#
#     # here we use the 1-tailed, since we are only interested in the case
#     # of the correlation coefficient being higher than the real one
#     corr_m_per_cell, corr_m_per_cell_pval, _ = get_mean_signif2(corr, corr_col)
#
#     # getting cdfs and pdfs
#     xmin = -1
#     xmax = 1
#     n_bins_cdf = 100
#     bins_cdf = np.linspace(xmin, xmax, n_bins_cdf + 1)
#
#     cdf_true = get_pdf_2(corr.values, bins_cdf, cdf_bool=True)
#     cdf_true = pd.DataFrame(cdf_true, index=cell_list)
#
#     cdfs_shfl = get_pdf_3(corr_col, bins_cdf, cdf_bool=True, checknan=False)
#
#     cdf_shfl_m = pd.DataFrame(cdfs_shfl.mean(axis=0), index=cell_list)
#
#     cdf_shfl_diff_max = get_min_diff(cdfs_shfl, cdf_shfl_m.values)
#     # the max is over the cdfs, i.e., the last dimension
#
#     # first dim are the cells, second dim are the odors
#     cdf_diff_max = get_min_diff(cdf_true, cdf_shfl_m)
#     # cdf_diff_max_true contains the true max deviation for each cell
#     # cdf_diff_max_true is n_cells
#
#     pvals = np.mean(cdf_shfl_diff_max >= cdf_diff_max.values, axis=0)
#     cdf_diff_max_pval = pd.Series(pvals, index=cell_list)
#
#     return (corr, corr_pval, corr_m_per_cell, corr_m_per_cell_pval,
#             cdf_diff_max, cdf_diff_max_pval, cdf_true, cdfs_shfl)


# #############################################################################
# ###############  FUNCTIONS FINDING BEST SUBSPACE PROJECTIONS  ###############
# #############################################################################
# functions finding the projectinos onto subspaces
def get_coef_proj_vect_on_subspace(v, subspace):
    """
    subspace is a collection of column vectors which form a subspace
    v is a vector
    find funtion output the coefficients a so that subspace.dot(a)
    is the the projection of v onto the subspace
    this vector is the one that has the smallest angle with v in the subspace
    """
    Q, R = LA.qr(subspace, mode='economic')
    a = LA.inv(R).dot(Q.T.dot(v))  # the coefficients for combining
    # the vectors of the subspace
    return a


def get_coef_proj_vect_on_subspace_c(v, subspace):
    """
    Does the same as the one above, it projects onto the
    subspace formed by the centered vectors
    """
    # v_cn = FG.ctr_norm(v)
    subspace_c = subspace - subspace.mean()  # center the vects in the subspace
    a = get_coef_proj_vect_on_subspace(v, subspace_c)
    # normally we are looking at the best projection of the center-normalized
    # vector v. Subtracting the mean of a vector is the same as doing an ortho-
    # gonal projection of that vector onto that space, that's why we don't
    # need to do explicitely subtract the mean of v before the operation above,
    # and multiplying by Q^T is enough
    return a


def get_cc_a_subspace(v, subspace, i=None):
    """
    Gives the highest correlation coefficient between a vector
    and a subspace formed by several vectors, as we as the coefficients
    to of the best vector in terms of the original vectors
    """
    a = get_coef_proj_vect_on_subspace_c(v, subspace.iloc[:, 0:i])

    # the vector with the highest corr. coef that is part of the given subspace
    v_proj = subspace.iloc[:, 0:i].dot(a)
    return np.corrcoef(v, v_proj)[0, 1], a


def get_cc_subspace_qr(v, subspace, i=None):
    """
    Gives the highest correlation coefficient between a vector
    and a subspace formed by several vectors, as we as the coefficients
    to of the best vector in terms of the original vectors
    # this function is the more direct than the one above, i.e., does less
    calculations to get to the same result
    """
    v_cn = get_ctr_norm(v)
    subspace_c = subspace - subspace.mean()
    Q, _ = LA.qr(subspace_c, mode='economic')
    Q1 = Q[:, 0:i]
    P = Q1.dot(Q1.T)
    proj = P @ v_cn
    proj_cn = get_norm(proj)

    return v_cn @ proj_cn


def get_cc_subspace_inv(v, subspace, i=None):
    """
    Gives the highest correlation coefficient between a vector
    and a subspace formed by several vectors, as we as the coefficients
    to of the best vector in terms of the original vectors

    this version of the function is wice quicker than the one doing
    the QR factorization.
    """
    v_cn = get_ctr_norm(v)
    A = subspace.values[:, 0:i]
    B = A - A.mean()
    # print(B.shape)
    P = B.dot(LA.inv(B.T.dot(B))).dot(B.T)
    proj = P @ v_cn
    proj_cn = get_norm(proj)

    return v_cn @ proj_cn


def get_cc_subspaces(v, subspace):
    """
    give all the correlation coefficients for all the iteratively created
    subspaces based on the columns vectors in subspace
    v is usually the connection vector
    """
    v_cn = get_ctr_norm(v)
    subspace_c = subspace - subspace.mean()
    projs = np.zeros(subspace.shape)
    Q, _ = LA.qr(subspace_c, mode='economic')
    n = subspace.shape[1]
    for i in range(n):
        Q1 = Q[:, 0:i+1]
        P = Q1.dot(Q1.T)
        if P.ndim != 2:
            raise ValueError(f'P is not as if should be: {P}')
        projs[:, i] = P @ v_cn

    projs_cn = get_norm_np(projs)
    return v_cn @ projs_cn

# function replaced by the ones below
# =============================================================================
# def get_cc_subspaces2(vs, subspace):
#     """
#     give all the correlation coefficients for all the iteratively created
#     subspaces based on the columns vectors in subspace
#     vs are vectors: d x n_cells
#     subspace: d x n_ss
#
#     projs:  d x n_cells x n_ss
#     Q: d x n_ss
#     P: d x d
#     P @ vs : d x n_cells
#
#     projs_cn: n_ss x d x n_cells
#     """
#     if subspace.shape[0] != vs.shape[0]:
#         raise ValueError("subspace and vs don't have the same number of rows"
#                          f'dimensions are {subspace.shape} and {vs.shape}')
#     vs_cn = get_ctr_norm(vs)
#     subspace_c = subspace - subspace.mean()
#     projs = np.zeros((vs.shape[0], vs.shape[1], subspace.shape[1]))
#     # projs is a 3d dataset, the dimensions are:
#     # n_columns in vs
#     Q, _ = LA.qr(subspace_c, mode='economic')
#     n = subspace.shape[1]
#     for i in range(n):
#         Q1 = Q[:, 0:i+1]
#         P = Q1.dot(Q1.T)
#         if P.ndim != 2:
#             raise ValueError(f'P is not as if should be: {P}')
#         projs[:, :, i] = P @ vs
#
#     projs_cn = get_norm_np(projs)
#     # print(projs_cn.shape)
#     projs_cn = np.moveaxis(projs_cn, -1, 0)
#     # the last axis become the 1st, this is necessary for the multiplication
#     # print(projs_cn.shape, vs_cn.shape)
#     ccs = np.sum(np.multiply(projs_cn, vs_cn.values), axis=1)
#     # first the multiplication make an element by element multiplication
#     # the sum gives us the correlation coefficient, summed over the axis 1,
#     # which is the axis of ORNs
#     return ccs
#
# =============================================================================


def create_proj_matrices(subspace):
    """
    creates projection matrices
    """
    (d, n) = subspace.shape
    Ps = np.zeros((n, d, d))
    Q, _ = LA.qr(subspace, mode='economic')
    for i in range(n):
        Q1 = Q[:, 0:i+1]
        Ps[i] = Q1.dot(Q1.T)
        if Ps[i].ndim != 2:
            raise ValueError(f'Ps is not as if should be: {Ps}')
    return Ps


def get_cc_subspaces2_P(vs, Ps):
    """
    give all the correlation coefficients for all the projection matrices
    vs are vectors: d x n_cells. vs needs to be either just centered, or
    centered and normalized, depending if you do cos sim or corr ceof
    Ps: n_ss x d x d

    projs:  d x n_cells x n_ss
    Ps[i] @ vs : d x n_cells

    projs_cn: n_ss x d x n_cells
    """
    if not is_norm_1_np(vs):
        raise ValueError('dataset not normalized: '
                         f'{np.abs(LA.norm(vs, axis=0)-1)}')

    if Ps.shape[1] != vs.shape[0]:
        raise ValueError("P and vs don't have the same number of rows, "
                         f'dimensions are {Ps.shape} and {vs.shape}')

    projs = np.zeros((vs.shape[0], vs.shape[1], Ps.shape[0]))
    # projs is a 3d dataset, the dimensions are:
    # n_columns in vs
    n = Ps.shape[0]
    for i in range(n):
        projs[:, :, i] = Ps[i] @ vs

    # if the Ps are already projecting on a centered subspace, then projs
    # don't need to be centered. If Ps project on the non-centered subspace
    # then also it does not need to be centered! However normalization is
    # necessary to calculate either the cos sim or the corr coef
    projs2 = get_norm_np(projs)
    # print(projs_cn.shape)
    projs2 = np.moveaxis(projs2, -1, 0)
    # the last axis become the first, this is necessary for the multiplication
    # print(projs_cn.shape, vs_cn.shape)
    ccs = np.sum(np.multiply(projs2, vs), axis=1)
    # first the multiplication make an element by element multiplication
    # the sum gives us the correlation coefficient, summed over the axis 1,
    # which is the axis of ORNs
    return ccs  # returs either CC or CS


def get_cc_subspcs2(subspace, vs, ctr=True):
    """
    give all the correlation coefficients for all the iteratively created
    subspaces based on the columns vectors in subspace
    vs are vectors: d x n_cells
    subspace: d x n_ss

    projs:  d x n_cells x n_ss
    Q: d x n_ss
    P: d x d
    P @ vs : d x n_cells

    projs_cn: n_ss x d x n_cells
    """
    if ctr:
        subspace2 = subspace - subspace.mean()
        vs2 = get_ctr_norm(vs).values
    else:
        subspace2 = subspace
        vs2 = get_norm(vs).values
    Ps = create_proj_matrices(subspace2)
    return get_cc_subspaces2_P(vs2, Ps)


def get_cc_pv_subspcs2(subspace, vs, N, ctr=True):
    """
    same as function above, plus the pvalues
    """
    if ctr:
        subspace2 = subspace - subspace.mean()
        vs2 = get_ctr_norm(vs).values
    else:
        subspace2 = subspace
        vs2 = get_norm(vs).values

    Ps = create_proj_matrices(subspace2)

    cc_true = get_cc_subspaces2_P(vs2, Ps)

    cc_shfl = np.zeros((N, *cc_true.shape))
    for i in range(N):
        cc_shfl[i] = get_cc_subspaces2_P(np.random.permutation(vs2), Ps)

    pvs = np.mean(cc_shfl >= cc_true, axis=0)

    cc_true = pd.DataFrame(cc_true, index=subspace.columns, columns=vs.columns)
    pvs = pd.DataFrame(pvs, index=subspace.columns, columns=vs.columns)

    return cc_true, pvs

# #############################################################################
# ################  FUNCTIONS COMPARING SUBSPACES  ############################
# #############################################################################


def subspc_overlap(A: np.array, B: np.array, scaled: int = 0) -> float:
    """
    compare subspaces
    With n1 and n2 the number of vectors in A and B, M = (n1+n2) is the
    maximum ('worse') norm, m = |n1-n2| is the minimum norm
    So if we want a normalized answer, between 0 and 1, an option is

    n = (N - m)/(M-m)
    and 1 - n gives 0 for not related and 1 for related, which
    is more similar to the correlation coefficient

    this scaling only works when n1 + n2 is smaller than the
    dimensionality of the space, which should be in most cases

    Parameters
    ----------
    A:
        matrix D x n2
    B:
        matrix D x n1
    scaled:
        flag, 0: no scaling, 1: divided by the max norm, 2: scaled between
        0 and 1

    Returns:
    --------
    err:
        the (relative) error
    """

    n1 = A.shape[1]
    n2 = B.shape[1]

    # making it quicker than the scipy function
    # norm = LA.norm(A @ A.T - B @ B.T)**2

    M = n1 + n2
    S = B.T.dot(A)
    norm = M - 2 * np.sum(S**2)

    if scaled == 0:
        pass
    elif scaled == 1:
        norm /= M
    else:
        m = np.abs(n1 - n2)
        norm = 1. - (norm - m)/(M-m)
    return norm


def get_subspc_overlap_fast(AAT: np.array, BBT: np.array, n1: int, n2: int)\
            -> float:
    """
    = compare subspaces
    returns a number between 0  and 1. 0 - no overlap, 1 - full overlap
    we assume that A and B are made of column vectors which are orthonormal
    inside the norm the dimension of the matrix will be the dimension of the
    space
    With n1 and n2 the number of vectors in A and B, M = (n1+n2) is the
    maximum ('worse') norm, m = |n1-n2| is the minimum norm
    So if we want a normalized answer, between 0 and 1, an option is

    n = (N - m)/(M-m)
    and 1 - n gives 0 for not related and 1 for related, which
    is more similar to the correlation coefficient

    Mitya might want to know some absolute numbers, you can do that later.

    Also, this scaling only works when n1 + n2 is smaller than the
    dimensionality of the space, which should be in most cases

    There might be an issue with the 1 dimensional case, will need to see.
    The input in the 1D case has to be in a form of a matrix, so 2 layers
    of brackets please

    function speed could be improved by using the faster matrix multiplications
    from BLAS and float32 instead of float64. See optimization done
    in the function calculating the SNMF_GD_offline
    """
    print('this function get_subspc_overlap_fast should not be used as it is '
          'actually much slower than the other version. RECODE!!!!')
    # if we want to work with the norm instead of the square of the norm
# =============================================================================
#     norm = LA.norm(A @ A.T - B @ B.T)
#     m = np.sqrt(np.abs(n1 - n2))
#     M = np.sqrt(n1 + n2)
# =============================================================================
    # making it quicker than the scipy function
    # norm = LA.norm(A @ A.T - B @ B.T)**2
    diff = AAT - BBT
    norm = np.sum(diff**2)
    m = np.abs(n1 - n2)
    M = n1 + n2
    # print(n1, n2, m, M, norm)

    # some other versions that i tried before:
    # LA.norm(con_Q - act_Q @ act_Q.T @ con_Q)

    return 1. - (norm - m)/(M-m)


def subspcs1_overlap(A, B):
    """
    compares sets of subspaces. The sets of subspaces are constructed by using
    the set of column vectors in A and then created subspaces by
    cumulatively adding the vectors. So the first subset is A[:, 0], the 2nd
    is made of the 2 vectors A[:, 0:2], etc...

    B is just taken vector by vector, no subspaces created with B

    A is the matrix from which we iteratively build up subspaces.
    usually the activity vector
    """
    n1 = len(A.T)
    n2 = len(B.T)
    ss_overlap = pd.DataFrame(index=A.columns, columns=B.columns, dtype=float)
    A_Q, _ = LA.qr(A, mode='economic')
    B_n = get_norm(B).values
    for i, j in itertools.product(range(n1), range(n2)):
        ss_overlap.iat[i, j] = subspc_overlap(A_Q[:, :i + 1], B_n[:, j:j + 1])
    # this way of extracting a vector from B keep the matrix structure
    return ss_overlap


def subspcs1_overlap_pv(A, B, n_perm=10):
    """
    compares sets of subspaces. The sets of subspaces are constructed by using
    the set of column vectors in A and then created subspaces by
    cumulatively adding the vectors. So the first subset is A[:, 0], the 2nd
    is made of the 2 vectors A[:, 0:2], etc...
    
    B is just taken vector by vector, no subspaces created with B, B is being
    shuffles
    
    A is the matrix from which we iteratively build up subspaces.
    usually the activity vector
    B is usually the connection vector
    """
    n1 = len(A.T)
    n2 = len(B.T)
    ss_overlap = pd.DataFrame(index=A.columns, columns=B.columns, dtype=float)
    ss_overlap_pv = pd.DataFrame(index=A.columns, columns=B.columns,
                                 dtype=float)

    A_Q, _ = LA.qr(A, mode='economic')

    # we are not computing subspaces from B, just taking the vector
    # however for the subspace overlap, we need the norm of the vectors in B
    B_n = get_norm(B).values
    for i in range(n1):
        for j in range(n2):
            ss_overlap.iat[i, j] = subspc_overlap(A_Q[:, :i + 1],
                                                       B_n[:, j:j+1])
        overlaps = np.zeros((n_perm, n2), dtype=float)
        for k in range(n_perm):
            B_sh = np.random.permutation(B_n)
            for j in range(n2):
                overlaps[k, j] = subspc_overlap(A_Q[:, :i + 1],
                                                    B_sh[:, j:j+1])
        # print(overlaps.shape, ss_overlap.iloc[i].shape)
        ss_overlap_pv.iloc[i] = np.mean(overlaps > ss_overlap.iloc[i].values,
                                        axis=0)

    return ss_overlap, ss_overlap_pv


def subspcs1_overlap_pv_fast(A, B, n_perm=10):
    """
    compares sets of subspaces. The sets of subspaces are constructed by using
    the set of column vectors in A and then created subspaces by
    cumulatively adding the vectors. So the first subset is A[:, 0], the 2nd
    is made of the 2 vectors A[:, 0:2], etc...

    B is just taken vector by vector, no subspaces created with B, B is being
    shuffles

    A is the matrix from which we iteratively build up subspaces.
    usually the activity vector
    B is usually the connection vector
    """
    n1 = len(A.T)
    n2 = len(B.T)
    # d = len(B)
    # my_arange = np.arange(d)

    ss_overlap = pd.DataFrame(index=A.columns, columns=B.columns,
                              dtype=float)
    ss_overlap_pv = pd.DataFrame(index=A.columns, columns=B.columns,
                                 dtype=float)

    A_Q, _ = LA.qr(A, mode='economic')

    AAT = np.zeros((n1, len(A), len(A)), dtype=float)
    for i in range(n1):
        AAT[i] = A_Q[:, :i+1] @ A_Q[:, :i+1].T

    B_n = get_norm(B).values
    BBT = np.zeros((n2, len(B), len(B)), dtype=float)
    for j in range(n2):
        BBT[j] = B_n[:, j:j+1] @ B_n[:, j:j+1].T

    # we are not computing subspaces from B, just taking the vector
    # however for the subspace overlap, we need the norm of the vectors in B

    for i in range(n1):
        for j in range(n2):
            ss_overlap.iat[i, j] = get_subspc_overlap_fast(AAT[i],
                                                           BBT[j], i+1, 1)
        overlaps = np.zeros((n_perm, n2), dtype=float)
        for k in range(n_perm):
            # doesn't seem that permutation is much faster than multiplication
            B_sh = np.random.permutation(B_n)
            # perm = np.random.permutation(my_arange)
            # BBT_shfl = BBT[:, perm][:, :, perm]
            for j in range(n2):
                overlaps[k, j] = get_subspc_overlap_fast(AAT[i],
                                                   # BBT_shfl[j], i+1, j+1)
                                    # np.outer(B_sh[:, j], B_sh[:, j]), n1, n2)
                                B_sh[:, j:j+1] @ B_sh[:, j:j+1].T, i+1, 1)
        # print(overlaps.shape, ss_overlap.iloc[i].shape)
        ss_overlap_pv.iloc[i] = np.mean(overlaps > ss_overlap.iloc[i].values,
                                        axis=0)

    return ss_overlap, ss_overlap_pv


def subspcs2_overlap(A, B):
    """
    compares sets of subspaces. The sets of subspaces are constructed by using
    the set of column vectors in A and B and then created subspaces by
    cumulatively adding the vectors. So the first subset is A[:, 0], the 2nd
    is made of the 2 vectors A[:, 0:2], etc...
    Same for B
    And then all the combinations of these subset are compared between each
    other, thus resturning a matrix as a result.
    comp stands for compare
    """
    n1 = len(A.T)
    n2 = len(B.T)
    ss_overlap = pd.DataFrame(index=A.columns, columns=B.columns, dtype=float)
    A_Q, _ = LA.qr(A, mode='economic')
    B_Q, _ = LA.qr(B, mode='economic')
    for i, j in itertools.product(range(n1), range(n2)):
        ss_overlap.iat[i, j] = subspc_overlap(A_Q[:, :i + 1], B_Q[:, :j + 1])

    return ss_overlap


def subspcs2_overlap_pv(A, B, n_perm=10):
    """
    A is being shuffled to calculate the significance
    there is a way to make this function faster, by always shuffling the
    whole A matrix, and then use parts of it to
    comp stands for compare

    normally the whole A is shuffled. However we could also just shuffle
    the last column, which will investigate if adding that vector is
    meaningful for the connection.
    """
    n1 = len(A.T)
    n2 = len(B.T)
    ss_overlap = pd.DataFrame(index=A.columns, columns=B.columns, dtype=float)
    ss_overlap_pv = pd.DataFrame(index=A.columns, columns=B.columns,
                                 dtype=float)

    A_Q, _ = LA.qr(A, mode='economic')
    B_Q, _ = LA.qr(B, mode='economic')
    for i, j in itertools.product(range(n1), range(n2)):
        A_sel = A.iloc[:, : i + 1]
        A_Q_sel = A_Q[:, :i+1]
        B_Q_sel = B_Q[:, :j+1]
        ss_overlap.iat[i, j] = subspc_overlap(A_Q_sel, B_Q_sel)
        overlaps = np.zeros(n_perm)
        for k in range(n_perm):
            A_sh = A_sel.values.copy()
            for m in range(len(A_sh.T)):
            # for m in [len(A_sh.T) - 1]:
                np.random.shuffle(A_sh[:, m])
            A_Q_shfl, _ = LA.qr(A_sh, mode='economic')

            # if shuffling both A and B
# =============================================================================
#             B_sh = B_sel.values.copy()
#             for m in range(len(B_sh.T)):
#                 np.random.shuffle(B_sh[:, m])
#             B_Q, _ = LA.qr(B_sh, mode='economic')
# =============================================================================

            overlaps[k] = subspc_overlap(A_Q_shfl, B_Q_sel)

        ss_overlap_pv.iat[i, j] = np.mean(overlaps > ss_overlap.iat[i, j])

    return ss_overlap, ss_overlap_pv


